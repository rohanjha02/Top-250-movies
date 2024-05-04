import csv, io
from django.shortcuts import render, redirect
from .forms import UploadFileForm
from .models import Movie
from django.contrib import messages
from openpyxl import load_workbook
from django.shortcuts import render

def upload_file_view(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            if file.name.endswith('.csv'):
                data_set = file.read().decode('UTF-8')
                io_string = io.StringIO(data_set)
                next(io_string)  # Skip the header row
                reader = csv.reader(io_string, delimiter='\t', quotechar='"')
            elif file.name.endswith('.xlsx'):
                workbook = load_workbook(filename=file)
                worksheet = workbook.active
                reader = worksheet.iter_rows(values_only=True)
                next(reader)  # Skip the header row

            for row in reader:
                if len(row) != 9:
                    messages.error(request, f"Row skipped due to incorrect number of columns: {row}")
                    continue
                try:
                    Movie.objects.create(
                        title=row[0],
                        year=int(row[1]),
                        duration=row[2],
                        rated_as=row[3],
                        genre=row[4],
                        director=row[5],
                        rating=row[6].strip('\"').split("/")[0].strip(),
                        stars=row[7],
                        streaming_on=row[8].strip('\"')
                    )
                    messages.success(request, f"Movie '{row[0]}' added successfully.")
                except Exception as e:
                    messages.error(request, f"Error processing row {row}: {e}")
            return redirect('upload_success')
    else:
        form = UploadFileForm()
    return render(request, 'upload.html', {'form': form})

def upload_success(request):
    return render(request, 'upload_success.html')

def base_view(request):
    return render(request, 'base.html')

